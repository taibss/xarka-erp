import io
from datetime import datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def _format_time(dt: Optional[datetime]) -> str:
    """Format a datetime to 'HH:MM AM/PM' string without timezone conversion."""
    if not dt:
        return '--'
    hour = dt.hour
    minute = dt.minute
    period = 'PM' if hour >= 12 else 'AM'
    display_hour = hour % 12
    if display_hour == 0:
        display_hour = 12
    return f"{display_hour:02d}:{minute:02d} {period}"


def _format_date(d) -> str:
    """Format a date to 'DD Mon YYYY' string."""
    if not d:
        return '--'
    return d.strftime('%d %b %Y')


def build_rows(records) -> List[Dict]:
    """Convert Attendance ORM rows into flat dicts for report generation.

    Args:
        records: List of Attendance ORM objects with employee relationship loaded.

    Returns:
        List of flat dicts with formatted strings.
    """
    rows = []
    for rec in records:
        emp = getattr(rec, 'employee', None)
        punch_in_time = _format_time(rec.punch_in)
        punch_out_time = _format_time(rec.punch_out)
        hours = f"{rec.hours_worked:.2f}" if rec.hours_worked is not None else '--'

        # If both punch times are present, override "absent" — machine data is source of truth
        status = (rec.status or '--').capitalize()
        if rec.punch_in and rec.punch_out and status.lower() == 'absent':
            status = 'Present'

        # Normalize "Weeklyoff present" (from eSSL device) to "Present"
        if status.lower() == 'weeklyoff present':
            status = 'Present'

        # Sunday always shows "Sunday" regardless of underlying status
        if rec.date and rec.date.weekday() == 6:
            status = 'Sunday'

        # Detect single-punch day (punch_in == punch_out)
        if rec.punch_in and rec.punch_out and rec.punch_in == rec.punch_out:
            status = 'Present (Incomplete)'

        # Catch-all: normalize any other raw device status to standard labels
        if status not in ('Present', 'Present (Incomplete)', 'Absent', 'Sunday'):
            if 'absent' in status.lower():
                status = 'Absent'
            else:
                status = 'Present'

        rows.append({
            'date': _format_date(rec.date),
            'employee': emp.name if emp else f'ID {rec.employee_id}',
            'punch_in': punch_in_time,
            'punch_out': punch_out_time,
            'hours': hours,
            'status': status,
            'source': (rec.source or 'manual').replace('_', ' ').title(),
        })
    return rows


def build_summary(rows: List[Dict]) -> Dict:
    """Compute summary statistics from the report rows.

    Args:
        rows: List of row dicts from build_rows().

    Returns:
        Dict with summary metrics.
    """
    sunday_days = sum(1 for r in rows if r.get('status') == 'Sunday')
    working_rows = [r for r in rows if r.get('status') != 'Sunday']
    total_days = len(working_rows)
    present_days = sum(1 for r in working_rows if 'Present' in r.get('status', ''))
    absent_days = sum(1 for r in working_rows if r.get('status') == 'Absent')
    incomplete_days = sum(1 for r in working_rows if r.get('status') == 'Present (Incomplete)')
    attendance_percent = round(present_days / total_days * 100, 1) if total_days > 0 else 0

    return {
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'incomplete_days': incomplete_days,
        'sunday_days': sunday_days,
        'attendance_percent': attendance_percent,
    }


def _employee_summary(name: str, emp_rows: List[Dict]) -> Dict:
    """Build summary stats scoped to a single employee's rows."""
    sunday_days = sum(1 for r in emp_rows if r.get('status') == 'Sunday')
    working_rows = [r for r in emp_rows if r.get('status') != 'Sunday']
    total_days = len(working_rows)
    present_days = sum(1 for r in working_rows if 'Present' in r.get('status', ''))
    absent_days = sum(1 for r in working_rows if r.get('status') == 'Absent')
    incomplete_days = sum(1 for r in working_rows if r.get('status') == 'Present (Incomplete)')
    attendance_percent = round(present_days / total_days * 100, 1) if total_days > 0 else 0

    return {
        'employee_name': name,
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'incomplete_days': incomplete_days,
        'sunday_days': sunday_days,
        'attendance_percent': attendance_percent,
    }


def build_per_employee_summary(rows: List[Dict]) -> List[Dict]:
    """Group rows by employee and compute per-employee summary stats.

    Returns a list of dicts, one per employee, sorted by name.
    """
    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        groups[r.get('employee', 'Unknown')].append(r)
    result = [_employee_summary(name, emp_rows) for name, emp_rows in groups.items()]
    result.sort(key=lambda x: x['employee_name'])
    return result


HEADERS = ['Date', 'Employee', 'Punch In', 'Punch Out', 'Hours', 'Status', 'Source']

DARK_BG = '1F2937'
WHITE_FONT = Font(color='FFFFFF', bold=True, size=11)
HEADER_FILL = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
HEADER_BORDER = Border(
    bottom=Side(style='thin', color='9CA3AF'),
)
THIN_BORDER = Border(
    bottom=Side(style='hair', color='D1D5DB'),
)
CELL_FONT = Font(size=10)


def generate_attendance_excel(rows: List[Dict], start_date: str, end_date: str, employee_label: str, summary=None) -> io.BytesIO:
    """Generate an Excel report of attendance data.

    Returns a BytesIO object containing the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1, value='Attendance Report')
    title_cell.font = Font(bold=True, size=16, color=DARK_BG)
    title_cell.alignment = Alignment(horizontal='left')
    ws.row_dimensions[1].height = 30

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    subtitle = f'{start_date}  to  {end_date}  |  {employee_label}'
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(size=11, color='6B7280')
    sub_cell.alignment = Alignment(horizontal='left')
    ws.row_dimensions[2].height = 22

    # Summary block
    current_row = 4
    if summary:
        if isinstance(summary, list):
            # Per-employee summary table
            emp_headers = ['Employee', 'Total Days', 'Present', 'Absent', 'Incomplete', 'Sundays', 'Attendance %']
            for col_idx, header in enumerate(emp_headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = WHITE_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = HEADER_BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            for emp_idx, emp_sum in enumerate(summary):
                emp_values = [
                    emp_sum['employee_name'],
                    emp_sum['total_days'],
                    emp_sum['present_days'],
                    emp_sum['absent_days'],
                    emp_sum['incomplete_days'],
                    emp_sum['sunday_days'],
                    f"{emp_sum['attendance_percent']}%",
                ]
                for col_idx, value in enumerate(emp_values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.font = CELL_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical='center')
                    if emp_idx % 2 == 1:
                        cell.fill = ALT_ROW_FILL
                current_row += 1

            current_row += 1  # Empty row after summary
        else:
            # Single combined summary (original layout)
            summary_labels = [
                ('Total Days', summary['total_days']),
                ('Present', summary['present_days']),
                ('Absent', summary['absent_days']),
                ('Incomplete Punches', summary['incomplete_days']),
                ('Sundays', summary['sunday_days']),
                ('Attendance %', f"{summary['attendance_percent']}%"),
            ]
            for label, value in summary_labels:
                label_cell = ws.cell(row=current_row, column=1, value=label)
                label_cell.font = WHITE_FONT
                label_cell.fill = HEADER_FILL
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                label_cell.border = HEADER_BORDER

                value_cell = ws.cell(row=current_row, column=2, value=value)
                value_cell.font = Font(bold=True, size=11, color=DARK_BG)
                value_cell.alignment = Alignment(horizontal='left', vertical='center')
                value_cell.border = THIN_BORDER

                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
                ws.row_dimensions[current_row].height = 24
                current_row += 1

            current_row += 1  # Empty row after summary

    # Header row
    header_row = current_row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = HEADER_BORDER
    ws.row_dimensions[header_row].height = 28

    # Data rows
    for row_idx, row_data in enumerate(rows):
        excel_row = header_row + 1 + row_idx
        values = [
            row_data['date'],
            row_data['employee'],
            row_data['punch_in'],
            row_data['punch_out'],
            row_data['hours'],
            row_data['status'],
            row_data['source'],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Column widths
    col_widths = [14, 22, 14, 14, 10, 12, 16]
    for idx, width in enumerate(col_widths):
        ws.column_dimensions[chr(65 + idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_attendance_pdf(rows: List[Dict], start_date: str, end_date: str, employee_label: str, summary=None) -> io.BytesIO:
    """Generate a PDF report of attendance data.

    Returns a BytesIO object containing the .pdf file.
    """
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles['Title']
    elements.append(Paragraph('Attendance Report', title_style))
    elements.append(Spacer(1, 6 * mm))

    # Subtitle
    subtitle_style = styles['Normal']
    subtitle_style.fontSize = 11
    subtitle_style.textColor = colors.HexColor('#6B7280')
    elements.append(Paragraph(f'{start_date}  to  {end_date}  |  {employee_label}', subtitle_style))
    elements.append(Spacer(1, 8 * mm))

    # Summary block
    if summary:
        dark_bg = colors.HexColor('#1F2937')

        if isinstance(summary, list):
            # Per-employee summary table
            emp_headers = ['Employee', 'Total Days', 'Present', 'Absent', 'Incomplete', 'Sundays', 'Attendance %']
            summary_data = [emp_headers]
            for emp_sum in summary:
                summary_data.append([
                    emp_sum['employee_name'],
                    emp_sum['total_days'],
                    emp_sum['present_days'],
                    emp_sum['absent_days'],
                    emp_sum['incomplete_days'],
                    emp_sum['sunday_days'],
                    f"{emp_sum['attendance_percent']}%",
                ])
            summary_col_widths = [40 * mm, 22 * mm, 20 * mm, 20 * mm, 22 * mm, 20 * mm, 28 * mm]
            summary_table = Table(summary_data, colWidths=summary_col_widths, repeatRows=1)

            summary_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), dark_bg),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            # Alternating row colors
            for i in range(1, len(summary_data)):
                if i % 2 == 0:
                    summary_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F3F4F6'))
        else:
            # Single combined summary (original layout)
            summary_labels = [
                ('Total Days', summary['total_days']),
                ('Present', summary['present_days']),
                ('Absent', summary['absent_days']),
                ('Incomplete Punches', summary['incomplete_days']),
                ('Sundays', summary['sunday_days']),
                ('Attendance %', f"{summary['attendance_percent']}%"),
            ]
            summary_data = [
                [label for label, _ in summary_labels],
                [str(val) for _, val in summary_labels],
            ]
            summary_col_widths = [30 * mm] * len(summary_labels)
            summary_table = Table(summary_data, colWidths=summary_col_widths)

            summary_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), dark_bg),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, 1), 10),
                ('BOTTOMPADDING', (0, 1), (-1, 1), 6),
                ('TOPPADDING', (0, 1), (-1, 1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])

        summary_table.setStyle(summary_style)
        elements.append(summary_table)
        elements.append(Spacer(1, 8 * mm))

    # Table data
    table_data = [HEADERS]
    for row_data in rows:
        table_data.append([
            row_data['date'],
            row_data['employee'],
            row_data['punch_in'],
            row_data['punch_out'],
            row_data['hours'],
            row_data['status'],
            row_data['source'],
        ])

    col_widths = [40 * mm, 50 * mm, 32 * mm, 32 * mm, 22 * mm, 30 * mm, 36 * mm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table styling
    dark_bg = colors.HexColor('#1F2937')
    alt_bg = colors.HexColor('#F3F4F6')
    grid_color = colors.HexColor('#D1D5DB')

    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), dark_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # Data
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, grid_color),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # Alternating row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), alt_bg)

    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)
    output.seek(0)
    return output
